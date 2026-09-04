"""Verification du noyau non graphique.

Lancement :  python tests/test_koboimp.py
Aucune dependance de test externe : le script s'execute avec le seul venv du projet.
"""

import os
import sys
import tempfile
import threading
import xml.etree.ElementTree as ET
from datetime import date, datetime, time

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd  # noqa: E402

from koboimp import config as config_mod  # noqa: E402
from koboimp import engine as engine_mod  # noqa: E402
from koboimp import excel, kobo_api, registry, schema, security, validation, xmlbuild  # noqa: E402

PASSED = []
FAILED = []


def check(label, condition, detail=""):
    if condition:
        PASSED.append(label)
    else:
        FAILED.append(f"{label} :: {detail}")


def section(title):
    print(f"\n--- {title} ---")


# ==========================================================================
# Point 1 : noms de colonnes
# ==========================================================================
section("Point 1 - noms de balises")

check("nom simple valide", xmlbuild.is_valid_xml_name("nom_village"))
check("espace rejete", not xmlbuild.is_valid_xml_name("Nom du village"))
check("accent rejete", not xmlbuild.is_valid_xml_name("Age_moyen_eleve\u00e9"))
check("chiffre initial rejete", not xmlbuild.is_valid_xml_name("2024_total"))
check("deux-points rejete", not xmlbuild.is_valid_xml_name("Unnamed: 3"))
check("chemin de groupe valide", xmlbuild.is_valid_path("identification/nom_village"))

check("suggestion espace", xmlbuild.sanitize_xml_name("Nom du village") == "Nom_du_village",
      xmlbuild.sanitize_xml_name("Nom du village"))
check("suggestion accent", xmlbuild.sanitize_xml_name("\u00c2ge (ans)") == "Age_ans",
      xmlbuild.sanitize_xml_name("\u00c2ge (ans)"))
check("suggestion chiffre initial", xmlbuild.sanitize_xml_name("2024_total") == "_2024_total",
      xmlbuild.sanitize_xml_name("2024_total"))

valides, problemes = xmlbuild.inspect_columns(
    ["nom", "Nom du village", "Unnamed: 0", "nom", "groupe/age"]
)
check("colonnes valides retenues", valides == ["nom", "groupe/age"], str(valides))
check("3 problemes detectes", len(problemes) == 3, str(problemes))
check("doublon signale", any("double" in p["reason"] for p in problemes))


# ==========================================================================
# Point 2 : entiers relus en float
# ==========================================================================
section("Point 2 - entiers")

check("1.0 -> 1", xmlbuild.format_value(1.0) == "1", xmlbuild.format_value(1.0))
check("telephone 90941410.0", xmlbuild.format_value(90941410.0) == "90941410",
      xmlbuild.format_value(90941410.0))
check("decimal preserve", xmlbuild.format_value(1.5) == "1.5", xmlbuild.format_value(1.5))
check("entier natif", xmlbuild.format_value(42) == "42")
check("integer type force", xmlbuild.format_value(3.0, "integer") == "3",
      xmlbuild.format_value(3.0, "integer"))
check("booleen", xmlbuild.format_value(True) == "true")

serie = pd.Series([1, 2, None], dtype="object")
frame = pd.DataFrame({"n": [1.0, 2.0, float("nan")]})
check("colonne float pandas -> entiers",
      [xmlbuild.format_value(v) for v in frame["n"]] == ["1", "2", None],
      str([xmlbuild.format_value(v) for v in frame["n"]]))


# ==========================================================================
# Point 3 : cellules vides omises
# ==========================================================================
section("Point 3 - cellules vides")

check("None omis", xmlbuild.format_value(None) is None)
check("NaN omis", xmlbuild.format_value(float("nan")) is None)
check("chaine vide omise", xmlbuild.format_value("") is None)
check("espaces omis", xmlbuild.format_value("   ") is None)
check("NaT omis", xmlbuild.format_value(pd.NaT) is None)

paires = xmlbuild.row_to_pairs({"a": "x", "b": None, "c": "", "d": 0})
check("seules les valeurs presentes", paires == [("a", "x"), ("d", "0")], str(paires))
check("zero conserve", ("d", "0") in paires)


# ==========================================================================
# Point 6 : groupes imbriques
# ==========================================================================
section("Point 6 - groupes")

payload, instance = xmlbuild.build_submission_xml(
    [
        ("identification/nom", "Issa"),
        ("identification/village", "Say"),
        ("mesures/anthropo/poids", "12.4"),
        ("simple", "ok"),
    ],
    root_name="aBnqy3JN9LSGJEAHfMuieE",
    form_version="v1",
)
racine = ET.fromstring(payload)
check("racine = uid", racine.tag == "aBnqy3JN9LSGJEAHfMuieE")
check("attribut id", racine.get("id") == "aBnqy3JN9LSGJEAHfMuieE")
check("groupe cree une fois", len(racine.findall("identification")) == 1)
check("deux enfants dans le groupe", len(racine.find("identification")) == 2)
check("groupe imbrique", racine.find("mesures/anthropo/poids") is not None)
check("valeur imbriquee", racine.find("mesures/anthropo/poids").text == "12.4")
check("champ a plat", racine.find("simple").text == "ok")
check("instanceID present", racine.find("meta/instanceID").text == instance)
check("instanceID format uuid", instance.startswith("uuid:"))
check("version presente", racine.find("__version__").text == "v1")
check("formhub present", racine.find("formhub/uuid") is not None)

try:
    xmlbuild.build_submission_xml([("a", "b")], root_name="2mauvais", form_version="v1")
    check("uid invalide rejete", False, "aucune exception")
except ValueError:
    check("uid invalide rejete", True)


# ==========================================================================
# Point 7 : formats
# ==========================================================================
section("Point 7 - formats de valeurs")

check("date", xmlbuild.format_value(datetime(2026, 3, 4), "date") == "2026-03-04",
      xmlbuild.format_value(datetime(2026, 3, 4), "date"))
check("date depuis texte", xmlbuild.format_value("04/03/2026", "date") == "2026-03-04",
      xmlbuild.format_value("04/03/2026", "date"))
check("date depuis Timestamp",
      xmlbuild.format_value(pd.Timestamp("2026-03-04"), "date") == "2026-03-04")

dt = xmlbuild.format_value(datetime(2026, 3, 4, 14, 30, 5), "datetime")
check("datetime ISO", dt.startswith("2026-03-04T14:30:05.000"), dt)
check("datetime avec fuseau", dt[-6] in "+-" and dt[-3] == ":", dt)

heure = xmlbuild.format_value(time(14, 30, 5), "time")
check("time ISO", heure.startswith("14:30:05.000"), heure)
check("time avec fuseau", heure[-3] == ":", heure)

check("select_multiple virgules",
      xmlbuild.format_value("a,b,c", "select_multiple") == "a b c",
      xmlbuild.format_value("a,b,c", "select_multiple"))
check("select_multiple point-virgule",
      xmlbuild.format_value("a; b ;c", "select_multiple") == "a b c")
check("select_multiple deja espaces",
      xmlbuild.format_value("a b", "select_multiple") == "a b")
check("geopoint virgules",
      xmlbuild.format_value("13.51,2.11", "geopoint") == "13.51 2.11 0 0",
      xmlbuild.format_value("13.51,2.11", "geopoint"))
check("caracteres de controle retires",
      xmlbuild.format_value("bon\x0bjour") == "bonjour",
      xmlbuild.format_value("bon\x0bjour"))
check("colonne reservee ignoree",
      xmlbuild.row_to_pairs({"submitted": "1", "_motif": "x", "nom": "a"}) == [("nom", "a")])


# ==========================================================================
# Point 8 / 9 / 10 : schema, modele, correspondance
# ==========================================================================
section("Points 8-10 - schema et validation")

ASSET = {
    "uid": "aTESTFORM123",
    "name": "Enquete menage",
    "has_deployment": True,
    "deployment__active": True,
    "deployment__submission_count": 17,
    "deployed_version_id": "vDEPLOY9",
    "version_id": "vBROUILLON",
    "content": {
        "survey": [
            {"type": "start", "name": "start"},
            {"type": "text", "name": "nom", "label": ["Nom du chef de menage"], "required": True},
            {"type": "integer", "name": "age", "label": ["Age"], "required": False},
            {"type": "begin_group", "name": "localisation", "label": ["Localisation"]},
            {"type": "select_one", "select_from_list_name": "regions",
             "name": "region", "label": ["Region"], "required": True},
            {"type": "geopoint", "name": "gps", "label": ["Position"]},
            {"type": "end_group"},
            {"type": "select_multiple", "select_from_list_name": "cultures",
             "name": "cultures_pratiquees", "label": ["Cultures"]},
            {"type": "date", "name": "date_visite", "label": ["Date de visite"]},
            {"type": "image", "name": "photo", "label": ["Photo"]},
            {"type": "note", "name": "info", "label": ["Merci"]},
            {"type": "begin_repeat", "name": "membres", "label": ["Membres"]},
            {"type": "text", "name": "prenom", "label": ["Prenom"]},
            {"type": "end_repeat"},
        ],
        "choices": [
            {"list_name": "regions", "name": "niamey", "label": ["Niamey"]},
            {"list_name": "regions", "name": "tahoua", "label": ["Tahoua"]},
            {"list_name": "cultures", "name": "mil", "label": ["Mil"]},
            {"list_name": "cultures", "name": "sorgho", "label": ["Sorgho"]},
        ],
    },
}

forme = schema.parse_asset(ASSET)
check("uid lu", forme.uid == "aTESTFORM123")
check("titre lu", forme.title == "Enquete menage")
check("version deployee preferee", forme.version == "vDEPLOY9", forme.version)
check("deploiement actif", forme.deployed)
check("note exclue", forme.get("info") is None)
check("chemin de groupe", forme.get("localisation/region") is not None)
check("acces par nom court", forme.get("region") is not None)
check("type select_one", forme.get("localisation/region").type == "select_one")
check("choix rattaches", forme.get("localisation/region").choice_names == {"niamey", "tahoua"})
check("repeat detecte", forme.has_repeats)
check("question de repeat marquee", forme.get("membres/prenom").in_repeat)
check("piece jointe detectee", forme.has_attachments)
check("importable exclut repeat et fichier",
      {q.path for q in forme.importable} == {
          "start", "nom", "age", "localisation/region", "localisation/gps",
          "cultures_pratiquees", "date_visite"},
      str(sorted(q.path for q in forme.importable)))
check("obligatoires", set(forme.required_paths) == {"nom", "localisation/region"},
      str(forme.required_paths))

DONNEES = pd.DataFrame({
    "nom": ["Issa", "Fati", None, "Ali"],
    "age": [34, "quarante", 12.0, 50],
    "localisation/region": ["niamey", "tahoua", "niamey", "kollo"],
    "cultures_pratiquees": ["mil,sorgho", "mil", "banane", None],
    "date_visite": ["2026-01-05", "pas une date", "2026-02-01", "2026-02-02"],
    "Nom du village": ["Say", "Say", "Say", "Say"],
    "colonne_inconnue": [1, 2, 3, 4],
})

rapport = validation.validate_dataframe(DONNEES, forme)
etats = {status.column: status.status for status in rapport.columns}
check("colonne reconnue", etats["nom"] == validation.COL_OK)
check("colonne de groupe reconnue", etats["localisation/region"] == validation.COL_OK)
check("nom invalide detecte", etats["Nom du village"] == validation.COL_INVALID)
check("colonne inconnue detectee", etats["colonne_inconnue"] == validation.COL_UNKNOWN)
check("suggestion fournie",
      any(s.suggestion == "Nom_du_village" for s in rapport.columns))
check("pas d'obligatoire manquant", rapport.missing_required == [], str(rapport.missing_required))

motifs = {(i.row_number, i.column) for i in rapport.row_issues}
check("obligatoire vide detecte", (4, "nom") in motifs, str(sorted(motifs)))
check("entier invalide detecte", (3, "age") in motifs, str(sorted(motifs)))
check("choix hors liste detecte", (5, "localisation/region") in motifs, str(sorted(motifs)))
check("choix multiple hors liste", (4, "cultures_pratiquees") in motifs, str(sorted(motifs)))
check("date illisible detectee", (3, "date_visite") in motifs, str(sorted(motifs)))
check("lignes invalides comptees", rapport.invalid_rows == 3, str(rapport.invalid_rows))
check("lignes valides comptees", rapport.valid_rows == 1, str(rapport.valid_rows))
check("avertissement repeat", any("repete" in w for w in rapport.warnings), str(rapport.warnings))
check("avertissement piece jointe", any("photos" in w for w in rapport.warnings))

manquant = validation.validate_dataframe(DONNEES.drop(columns=["nom"]), forme)
check("obligatoire manquant bloque", manquant.has_blocking_errors)
check("obligatoire manquant liste", [q.name for q in manquant.missing_required] == ["nom"])

ok, motif = validation.validate_row_values({"nom": "Issa", "localisation/region": "niamey"}, forme)
check("point 24 ligne valide", ok, motif)
ok, motif = validation.validate_row_values({"nom": "", "localisation/region": "niamey"}, forme)
check("point 24 ligne invalide", not ok and "obligatoire" in motif, motif)


# ==========================================================================
# Point 4 / 14 / 20 : reponses serveur
# ==========================================================================
section("Points 4, 14, 20 - reponses serveur")


class FausseReponse:
    def __init__(self, status_code, text="", headers=None):
        self.status_code = status_code
        self.text = text
        self.headers = headers or {}


class FausseSession:
    def __init__(self, reponses):
        self.reponses = list(reponses)
        self.appels = 0
        self.headers = {}

    def post(self, *_args, **_kwargs):
        self.appels += 1
        reponse = self.reponses.pop(0)
        if isinstance(reponse, Exception):
            raise reponse
        return reponse

    def close(self):
        pass


def client_avec(reponses, **kwargs):
    client = kobo_api.KoboClient(token="t", submission_base_url="https://x", **kwargs)
    session = FausseSession(reponses)
    client._local.session = session
    return client, session


client, session = client_avec([FausseReponse(201)])
res = client.submit(b"<x/>")
check("201 = succes", res.status == kobo_api.SUCCESS and res.sent)

client, session = client_avec([FausseReponse(202, "Duplicate submission")])
res = client.submit(b"<x/>")
check("202 = doublon accepte", res.status == kobo_api.DUPLICATE and res.sent, res.status)
check("202 sans reessai", session.appels == 1, str(session.appels))

client, session = client_avec([FausseReponse(400, "<OpenRosaResponse xmlns='http://openrosa.org/http/response'>"
                                                  "<message nature='submit_error'>Invalid value for age</message>"
                                                  "</OpenRosaResponse>")])
res = client.submit(b"<x/>")
check("400 = refus", res.status == kobo_api.REJECTED)
check("400 sans reessai", session.appels == 1, str(session.appels))
check("message OpenRosa extrait", "Invalid value for age" in res.message, res.message)
check("400 explique en francais", "refusees" in res.message.lower(), res.message)

client, session = client_avec([FausseReponse(401, '{"detail":"Invalid token."}')])
res = client.submit(b"<x/>")
check("401 = authentification", res.status == kobo_api.AUTH)
check("401 sans reessai", session.appels == 1)
check("detail JSON extrait", "Invalid token." in res.message, res.message)

client, session = client_avec(
    [FausseReponse(500), FausseReponse(500), FausseReponse(201)], max_attempts=4
)
arret = threading.Event()
kobo_api.BACKOFF_BASE = 1.0  # accelere le test
res = client.submit(b"<x/>", stop_event=arret)
check("5xx reessaye puis reussit", res.status == kobo_api.SUCCESS, res.status)
check("trois tentatives", session.appels == 3, str(session.appels))

client, session = client_avec([FausseReponse(429, "", {"Retry-After": "1"}), FausseReponse(201)],
                              max_attempts=3)
res = client.submit(b"<x/>", stop_event=threading.Event())
check("429 respecte Retry-After puis reussit", res.status == kobo_api.SUCCESS)

import requests as _requests  # noqa: E402
client, session = client_avec(
    [_requests.exceptions.ConnectionError("coupure"), FausseReponse(201)], max_attempts=3
)
res = client.submit(b"<x/>", stop_event=threading.Event())
check("coupure reseau reessayee", res.status == kobo_api.SUCCESS, res.status)

client, session = client_avec([FausseReponse(500)] * 5, max_attempts=3)
arret = threading.Event()
arret.set()
res = client.submit(b"<x/>", stop_event=arret)
check("arret immediat", res.status == kobo_api.STOPPED, res.status)
check("aucun envoi apres arret", session.appels == 0, str(session.appels))

# --- endpoint de soumission (API v1 supprimee par Kobo en juin 2026) ---------
check("endpoint OpenRosa", kobo_api.SUBMISSION_PATH == "/submission",
      kobo_api.SUBMISSION_PATH)
client, session = client_avec([FausseReponse(201)])
check("URL d'envoi complete",
      client.submission_url() == "https://x/submission", client.submission_url())

# Un 410 est definitif : le reessayer quatre fois par ligne n'a aucun sens.
client, session = client_avec(
    [FausseReponse(410, "<html><body>Gone</body></html>")] * 5, max_attempts=4
)
res = client.submit(b"<x/>", stop_event=threading.Event())
check("410 = endpoint supprime", res.status == kobo_api.GONE, res.status)
check("410 sans reessai", session.appels == 1, str(session.appels))
check("410 non reprenable comme panne", not res.retryable_later)
check("410 explique la marche a suivre",
      "supprimee" in res.message and "Mettez a jour" in res.message, res.message)
check("410 compte comme echec reprenable",
      "GONE" in registry.FAILURE_STATUSES)

# Avec un hote de repli connu, on tente l'autre serveur avant de conclure.
client = kobo_api.KoboClient(token="t", submission_base_url="https://kf.test",
                             fallback_submission_base_url="https://kc.test",
                             max_attempts=4)
session = FausseSession([FausseReponse(410), FausseReponse(201)])
client._local.session = session
res = client.submit(b"<x/>", stop_event=threading.Event())
check("410 declenche le repli d'hote", res.status == kobo_api.SUCCESS, res.status)
check("repli effectue en 2 appels", session.appels == 2, str(session.appels))
check("hote de repli retenu",
      client.submission_url() == "https://kc.test/submission", client.submission_url())

check("message 404 en francais",
      "introuvable" in kobo_api.humanize_http_error(404, "").lower())
check("html nettoye",
      "<" not in kobo_api.humanize_http_error(500, "<html><body>Boom</body></html>"))


# ==========================================================================
# Point 5 / 21 : registre
# ==========================================================================
section("Points 5 et 21 - registre")

with tempfile.TemporaryDirectory() as dossier:
    base = registry.Registry(os.path.join(dossier, "reg.db"))

    table = pd.DataFrame({"a": [1, 2, 3, 2], "b": ["x", "y", "z", "y"]})
    cles = registry.compute_row_keys(table, "SRC1")
    check("une cle par ligne", len(cles) == 4)
    check("lignes identiques distinguees", cles[1] != cles[3], f"{cles[1]} / {cles[3]}")
    check("cle stable entre appels", registry.compute_row_keys(table, "SRC1") == cles)

    melange = table.iloc[[2, 0, 1, 3]].reset_index(drop=True)
    cles_melange = registry.compute_row_keys(melange, "SRC1")
    check("cle insensible a l'ordre", set(cles_melange) == set(cles))

    etats = base.register_rows("SRC1", "fichier.xlsx", "aFORM", cles)
    check("4 lignes enregistrees", len(etats) == 4)
    check("statut initial", all(e["status"] == registry.PENDING for e in etats.values()))
    check("instanceID unique", len({e["instance_id"] for e in etats.values()}) == 4)

    premier = etats[cles[0]]["instance_id"]
    etats2 = base.register_rows("SRC1", "fichier.xlsx", "aFORM", cles)
    check("instanceID stable au reenregistrement", etats2[cles[0]]["instance_id"] == premier)

    base.mark_many([
        (cles[0], registry.SUCCESS, 201, "ok"),
        (cles[1], "REJECTED", 400, "valeur invalide"),
        (cles[2], "NETWORK", None, "coupure"),
    ])

    etats = base.states(cles)
    choisies, ignorees = registry.select_rows_to_send(etats, cles, "new")
    check("mode new ignore l'envoye", ignorees == 1 and choisies == [1, 2, 3],
          f"{ignorees} / {choisies}")

    choisies, ignorees = registry.select_rows_to_send(etats, cles, "retry")
    check("mode retry ne prend que les echecs", choisies == [1, 2], str(choisies))

    choisies, ignorees = registry.select_rows_to_send(etats, cles, "force")
    check("mode force prend tout", choisies == [0, 1, 2, 3] and ignorees == 0)

    resume = base.summary("SRC1")
    check("resume envoye", resume["_sent"] == 1, str(resume))
    check("resume echecs", resume["_failed"] == 2, str(resume))
    check("resume total", resume["_total"] == 4, str(resume))

    run = base.start_run("SRC1", "fichier.xlsx", "aFORM", "Titre", 4)
    base.finish_run(run, 1, 2, 1, False)
    check("execution historisee", base.recent_runs()[0]["sent"] == 1)

    check("oubli du fichier", base.forget_source("SRC1") == 4)
    check("registre vide apres oubli", base.summary("SRC1")["_total"] == 0)
    base.close()


# ==========================================================================
# Point 9 / 11 / 14 : Excel
# ==========================================================================
section("Points 9, 11, 14 - Excel")

with tempfile.TemporaryDirectory() as dossier:
    modele = os.path.join(dossier, "modele.xlsx")
    excel.build_template(forme, modele)
    check("modele cree", os.path.exists(modele))

    from openpyxl import load_workbook
    classeur = load_workbook(modele)
    check("feuille Donnees", excel.DATA_SHEET in classeur.sheetnames)
    check("feuille Notice", excel.NOTICE_SHEET in classeur.sheetnames)
    check("feuille Listes masquee", classeur[excel.LIST_SHEET].sheet_state == "hidden")

    entetes = [cell.value for cell in classeur[excel.DATA_SHEET][1]]
    check("entetes = chemins de questions", "localisation/region" in entetes, str(entetes))
    check("photo exclue du modele", "photo" not in entetes)
    check("question de repeat exclue", "membres/prenom" not in entetes)
    check("liste deroulante posee", len(classeur[excel.DATA_SHEET].data_validations.dataValidation) >= 1)
    classeur.close()

    lu, feuille = excel.read_table(modele, excel.DATA_SHEET)
    check("relecture du modele", list(lu.columns) == entetes, str(list(lu.columns)))
    check("feuille retournee", feuille == excel.DATA_SHEET)
    check("feuilles listees", excel.NOTICE_SHEET in excel.list_sheets(modele))

    rapport_path = os.path.join(dossier, "erreurs.xlsx")
    excel.write_error_report(
        rapport_path,
        DONNEES,
        [
            {"row_index": 1, "status": "REJECTED", "http_status": 400, "message": "age invalide"},
            {"row_index": 3, "status": "NETWORK", "http_status": None, "message": "coupure"},
        ],
        rapport,
        context={"Fichier source": "test.xlsx"},
    )
    check("rapport cree", os.path.exists(rapport_path))
    corrige = pd.read_excel(rapport_path, sheet_name="A corriger")
    check("2 lignes a corriger", len(corrige) == 2, str(len(corrige)))
    check("colonnes d'origine conservees", "nom" in corrige.columns, str(list(corrige.columns)))
    check("numero de ligne Excel", list(corrige[excel.REPORT_ROW_COLUMN]) == [3, 5],
          str(list(corrige[excel.REPORT_ROW_COLUMN])))
    check("motif present", "age invalide" in str(corrige[excel.REPORT_REASON_COLUMN].iloc[0]))
    check("feuille Details", "Details" in load_workbook(rapport_path).sheetnames)
    check("colonnes de rapport ignorees a la relecture",
          xmlbuild.row_to_pairs({excel.REPORT_ROW_COLUMN: 3, excel.REPORT_REASON_COLUMN: "x",
                                 "nom": "Issa"}) == [("nom", "Issa")])

    # Une ligne vide au milieu ne doit pas decaler la numerotation : sinon le
    # rapport designerait la mauvaise ligne a corriger.
    troue = os.path.join(dossier, "troue.xlsx")
    pd.DataFrame({
        "nom": ["Issa", None, "Ali", None, None],
        "age": [30, None, 41, None, None],
    }).to_excel(troue, index=False)
    lu, _ = excel.read_table(troue)
    check("lignes vides de fin retirees", len(lu) == 3, str(len(lu)))
    check("ligne vide du milieu conservee", pd.isna(lu["nom"].iloc[1]), str(list(lu["nom"])))
    check("numerotation Excel preservee", lu["nom"].iloc[2] == "Ali", str(list(lu["nom"])))

    # Un rapport corrige puis reimporte contient deja les colonnes de service :
    # elles doivent etre remplacees, pas dupliquees.
    regenere = os.path.join(dossier, "regenere.xlsx")
    deja_annote = pd.DataFrame({
        excel.REPORT_ROW_COLUMN: [3],
        excel.REPORT_REASON_COLUMN: ["ancien motif"],
        "nom": ["Fati"],
    })
    excel.write_error_report(regenere, deja_annote, [
        {"row_index": 0, "status": "REJECTED", "http_status": 400, "message": "toujours faux"},
    ])
    relu = pd.read_excel(regenere, sheet_name="A corriger")
    check("rapport regenerable depuis un rapport corrige", os.path.exists(regenere))
    check("colonnes de service non dupliquees",
          list(relu.columns).count(excel.REPORT_ROW_COLUMN) == 1, str(list(relu.columns)))
    check("motif mis a jour", "toujours faux" in str(relu[excel.REPORT_REASON_COLUMN].iloc[0]),
          str(relu[excel.REPORT_REASON_COLUMN].iloc[0]))

    manquant_path = os.path.join(dossier, "absent.xlsx")
    try:
        excel.read_table(manquant_path)
        check("fichier absent signale", False, "aucune exception")
    except excel.ExcelError as exc:
        check("fichier absent signale", "introuvable" in str(exc).lower(), str(exc))


# ==========================================================================
# Point 22 / 23 : chemins et token
# ==========================================================================
section("Points 22 et 23 - configuration")

jeton = "abcdef0123456789abcdef0123456789"
stocke = security.encrypt_token(jeton)
check("token transforme", stocke != jeton and stocke != "")
check("token relu", security.decrypt_token(stocke) == jeton)
check("token vide", security.encrypt_token("") == "" and security.decrypt_token("") == "")
check("ancien format en clair accepte", security.decrypt_token(jeton) == jeton)
if sys.platform == "win32":
    check("DPAPI utilisee", security.is_protected(stocke), stocke[:12])
check("masquage", security.mask(jeton).endswith("6789") and "*" in security.mask(jeton))

check("kc -> kf", config_mod.derive_kpi_url("https://kc.kobotoolbox.org") == "https://kf.kobotoolbox.org",
      config_mod.derive_kpi_url("https://kc.kobotoolbox.org"))
check("kf inchange", config_mod.derive_kpi_url("https://kf.kobotoolbox.org") == "https://kf.kobotoolbox.org")
check("eu inchange", config_mod.derive_kpi_url("https://eu.kobotoolbox.org") == "https://eu.kobotoolbox.org")
check("repli kc", config_mod.alternate_submission_url("https://kf.kobotoolbox.org") == "https://kc.kobotoolbox.org",
      config_mod.alternate_submission_url("https://kf.kobotoolbox.org"))
check("schema ajoute", config_mod.normalize_base_url("kf.kobotoolbox.org") == "https://kf.kobotoolbox.org")
check("barre finale retiree", config_mod.normalize_base_url("https://kf.kobotoolbox.org/") == "https://kf.kobotoolbox.org")

with tempfile.TemporaryDirectory() as dossier:
    ancien = {
        "server_base_url": "https://kc.kobotoolbox.org",
        "api_token": jeton,
        "assets_uid": "aVIEUX",
        "resume_only_failed": True,
        "success_dir": "C:/vieux/success",
        "failed_dir": "C:/vieux/failed",
        "max_workers": 3,
    }
    chemin = os.path.join(dossier, "vieux.json")
    import json
    with open(chemin, "w", encoding="utf-8") as handle:
        json.dump(ancien, handle)

    migre = config_mod.import_config(chemin)
    check("assets_uid migre", migre["asset_uid"] == "aVIEUX")
    check("token migre et chiffre", config_mod.get_token(migre) == jeton)
    check("token absent en clair", migre["api_token_enc"] != jeton)
    check("resume_only_failed migre", migre["resume_mode"] == "retry")
    check("dossiers obsoletes retires", "success_dir" not in migre and "failed_dir" not in migre)
    check("workers conserves", migre["max_workers"] == 3)

    export = os.path.join(dossier, "partage.json")
    config_mod.export_config(migre, export, include_token=False)
    with open(export, encoding="utf-8") as handle:
        partage = json.load(handle)
    check("export sans token", partage["api_token_enc"] == "")
    check("export conserve le reste", partage["asset_uid"] == "aVIEUX")


# ==========================================================================
# Points 16-19 : moteur de bout en bout
# ==========================================================================
section("Points 16-19 - moteur")


class ClientSimule:
    """Repond selon la valeur de la colonne nom, sans reseau."""

    def __init__(self):
        self.envois = []
        self.lock = threading.Lock()

    def submit(self, payload, filename="s.xml", stop_event=None):
        with self.lock:
            self.envois.append(payload)
        racine = ET.fromstring(payload)
        nom = racine.findtext("nom") or ""
        if nom == "Fati":
            return kobo_api.SubmitResult(kobo_api.REJECTED, 400, "age invalide", 1)
        if nom == "Ali":
            return kobo_api.SubmitResult(kobo_api.DUPLICATE, 202, "deja presente", 1)
        return kobo_api.SubmitResult(kobo_api.SUCCESS, 201, "ok", 1)


with tempfile.TemporaryDirectory() as dossier:
    table = pd.DataFrame({
        "nom": ["Issa", "Fati", "Ali", "Zara"],
        "age": [34.0, 22.0, 41.0, 29.0],
        "localisation/region": ["niamey", "niamey", "tahoua", "niamey"],
        "colonne_inconnue": ["a", "b", "c", "d"],
    })
    statuts = validation.map_columns(table.columns, forme)

    conf = {
        "dry_run": False,
        "resume_mode": "new",
        "max_workers": 3,
        "output_dir": os.path.join(dossier, "echecs"),
        "log_file": os.path.join(dossier, "journal.csv"),
        "report_dir": os.path.join(dossier, "rapports"),
    }

    base = registry.Registry(os.path.join(dossier, "reg.db"))
    client = ClientSimule()
    traces = []

    moteur = engine_mod.ImportEngine(
        conf, table, forme, statuts, "SRCX", "menages.xlsx", client, base,
        log_callback=lambda message, level="info": traces.append((level, message)),
        stop_event=threading.Event(),
        validation_report=None,
    )
    resultat = moteur.run()

    check("4 lignes traitees", resultat.processed == 4, str(resultat.processed))
    check("2 succes", resultat.sent == 2, str(resultat.sent))
    check("1 doublon compte a part", resultat.duplicates == 1, str(resultat.duplicates))
    check("1 echec", resultat.failed == 1, str(resultat.failed))
    check("4 envois emis", len(client.envois) == 4, str(len(client.envois)))

    envoi = ET.fromstring(client.envois[0])
    check("colonne inconnue exclue du XML", envoi.find("colonne_inconnue") is None)
    check("age converti en entier", envoi.findtext("age") in {"34", "22", "41", "29"},
          envoi.findtext("age"))
    check("groupe respecte dans l'envoi", envoi.find("localisation/region") is not None)

    check("journal CSV cree", os.path.exists(conf["log_file"]))
    with open(conf["log_file"], encoding="utf-8-sig") as handle:
        lignes = handle.read().strip().splitlines()
    check("entete + 4 lignes", len(lignes) == 5, str(len(lignes)))
    check("separateur point-virgule", lignes[0].count(";") == 4, lignes[0])

    echecs = os.listdir(conf["output_dir"]) if os.path.exists(conf["output_dir"]) else []
    check("point 16 : seul l'echec sur disque", len(echecs) == 1, str(echecs))
    check("nom de fichier d'echec explicite", echecs[0].startswith("echec_ligne_"), str(echecs))

    check("rapport genere", resultat.report_path and os.path.exists(resultat.report_path),
          resultat.report_path)
    corrige = pd.read_excel(resultat.report_path, sheet_name="A corriger")
    check("rapport contient la ligne fautive", list(corrige["nom"]) == ["Fati"], str(list(corrige["nom"])))

    # Relance : rien ne doit repartir (points 5 et 21).
    client2 = ClientSimule()
    moteur2 = engine_mod.ImportEngine(
        conf, table, forme, statuts, "SRCX", "menages.xlsx", client2, base,
        stop_event=threading.Event(),
    )
    resultat2 = moteur2.run()
    check("relance : 3 lignes deja passees ignorees", resultat2.skipped == 3, str(resultat2.skipped))
    check("relance : seul l'echec est retente", len(client2.envois) == 1, str(len(client2.envois)))

    # instanceID stable entre les deux executions (deduplication cote Kobo).
    id1 = ET.fromstring([p for p in client.envois if b"Fati" in p][0]).findtext("meta/instanceID")
    id2 = ET.fromstring(client2.envois[0]).findtext("meta/instanceID")
    check("instanceID stable au renvoi", id1 == id2, f"{id1} != {id2}")

    # Mode simulation : aucun envoi, XML sur disque.
    conf_sim = dict(conf, dry_run=True, resume_mode="force",
                    output_dir=os.path.join(dossier, "simu"))
    client3 = ClientSimule()
    moteur3 = engine_mod.ImportEngine(
        conf_sim, table, forme, statuts, "SRCX", "menages.xlsx", client3, base,
        stop_event=threading.Event(),
    )
    resultat3 = moteur3.run()
    check("simulation sans envoi", len(client3.envois) == 0)
    check("simulation ecrit les XML", len(os.listdir(conf_sim["output_dir"])) == 4,
          str(os.listdir(conf_sim["output_dir"])))

    # Une simulation ne doit pas effacer l'historique reel : sinon l'envoi
    # suivant recreerait des doublons pour des lignes deja arrivees.
    apres_simulation = base.summary("SRCX")
    check("simulation ne touche pas l'historique", apres_simulation["_sent"] == 3,
          str(apres_simulation))
    check("aucun statut DRY_RUN enregistre", "DRY_RUN" not in apres_simulation,
          str(apres_simulation))

    # Un journal CSV inaccessible ne doit pas empecher l'import.
    # Nom de fichier illegal sous Windows : provoque une vraie erreur d'ouverture.
    conf_bloque = dict(conf, resume_mode="force",
                       log_file=os.path.join(dossier, "jour<nal>.csv"))
    client_bloque = ClientSimule()
    traces_bloque = []
    moteur_bloque = engine_mod.ImportEngine(
        conf_bloque, table, forme, statuts, "SRC_LOG", "menages.xlsx", client_bloque, base,
        log_callback=lambda message, level="info": traces_bloque.append((level, message)),
        stop_event=threading.Event(),
    )
    resultat_bloque = moteur_bloque.run()
    check("journal illisible : import poursuivi", resultat_bloque.processed == 4,
          str(resultat_bloque.processed))
    check("journal illisible : avertissement emis",
          any("Journal CSV indisponible" in message for _, message in traces_bloque),
          str(traces_bloque[:3]))

    client4 = ClientSimule()
    moteur4 = engine_mod.ImportEngine(
        conf, table, forme, statuts, "SRCX", "menages.xlsx", client4, base,
        stop_event=threading.Event(),
    )
    resultat4 = moteur4.run()
    check("apres simulation, toujours un seul renvoi", len(client4.envois) == 1,
          str(len(client4.envois)))

    base.close()


# ==========================================================================
section("Resultat")
print(f"\n{len(PASSED)} verification(s) reussie(s), {len(FAILED)} echec(s).")
for echec in FAILED:
    print(f"  ECHEC  {echec}")
sys.exit(1 if FAILED else 0)
