"""Lecture des tableaux (Excel et CSV), modele telechargeable et rapport d'erreurs.

Points couverts :
   7 - le CSV est accepte en entree au meme titre qu'un classeur Excel : beaucoup
       d'exports tiers (ODK, KoboCollect, systemes metier) ne sortent que ce
       format, et passer par Excel pour le convertir introduit precisement les
       degradations que cette application corrige (entiers en 1.0, dates
       reinterpretees) ;
   9 - generation d'un modele Excel a partir du formulaire (bonnes en-tetes,
       types documentes, listes deroulantes pour les questions a choix) ;
  11 - rapport d'erreurs au format Excel, contenant les donnees d'origine et un
       motif en francais, directement reimportable apres correction ;
  14 - choix de la feuille et message clair quand le fichier est ouvert dans Excel.
"""

import csv
import hashlib
import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DATA_SHEET = "Donnees"
NOTICE_SHEET = "Notice"
LIST_SHEET = "Listes"

REPORT_ROW_COLUMN = "_ligne_excel"
REPORT_REASON_COLUMN = "_motif"

TEMPLATE_ROWS = 500

CSV_SUFFIXES = (".csv", ".txt", ".tsv")
EXCEL_SUFFIXES = (".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")

# Encodages essayes dans l'ordre. utf-8-sig en premier : c'est ce que produit
# Excel a l'enregistrement en CSV, et le BOM doit etre consomme, pas lu comme
# un caractere collé au premier en-tete. cp1252 ferme la marche pour les
# fichiers issus d'un Windows francais sans declaration d'encodage.
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_REQUIRED_FILL = PatternFill("solid", fgColor="B45309")
_TITLE_FILL = PatternFill("solid", fgColor="E7EDF3")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


class ExcelError(Exception):
    """Erreur de lecture presentable telle quelle a l'utilisateur."""


# --------------------------------------------------------------------------
# Lecture (point 14)
# --------------------------------------------------------------------------

def _friendly_error(path, exc):
    name = os.path.basename(path)
    if isinstance(exc, PermissionError):
        return ExcelError(
            f"Impossible de lire « {name} » : le fichier est ouvert dans Excel.\n\n"
            "Fermez-le puis relancez l'operation."
        )
    if isinstance(exc, FileNotFoundError):
        return ExcelError(f"Fichier introuvable : {path}")
    message = str(exc)
    if "not supported" in message.lower() or "zip" in message.lower():
        return ExcelError(
            f"« {name} » n'est pas un classeur Excel exploitable.\n\n"
            "Enregistrez-le au format .xlsx depuis Excel puis reessayez."
        )
    return ExcelError(f"Lecture impossible de « {name} » :\n{message}")


def _friendly_write_error(path, exc):
    name = os.path.basename(path)
    if isinstance(exc, PermissionError):
        return ExcelError(
            f"Impossible d'ecrire « {name} » : le fichier est deja ouvert dans Excel, "
            "ou le dossier est protege en ecriture.\n\n"
            "Fermez le fichier, ou choisissez un autre emplacement."
        )
    if isinstance(exc, (FileNotFoundError, NotADirectoryError)):
        return ExcelError(f"Dossier de destination introuvable :\n{os.path.dirname(path)}")
    return ExcelError(f"Ecriture impossible de « {name} » :\n{exc}")


def is_csv(path):
    return str(path or "").lower().endswith(CSV_SUFFIXES)


def _sniff_csv(path):
    """Devine encodage et separateur d'un CSV.

    Un CSV francais est souvent separe par des points-virgules (Excel suit la
    locale du poste) : imposer la virgule produirait une unique colonne portant
    toute la ligne, et un message d'erreur incomprehensible pour l'utilisateur.
    """
    for encoding in CSV_ENCODINGS:
        try:
            with open(path, "r", encoding=encoding, newline="") as handle:
                sample = handle.read(64 * 1024)
        except (UnicodeDecodeError, LookupError):
            continue
        except OSError as exc:
            raise _friendly_error(path, exc) from exc

        if not sample.strip():
            return encoding, ","

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,	|")
            return encoding, dialect.delimiter
        except csv.Error:
            # Sniffer echoue sur une seule colonne : on tranche au comptage.
            first = sample.splitlines()[0] if sample.splitlines() else ""
            counts = {sep: first.count(sep) for sep in (";", ",", "	", "|")}
            best = max(counts, key=counts.get)
            return encoding, best if counts[best] else ","

    raise ExcelError(
        f"Impossible de determiner l'encodage de « {os.path.basename(path)} ».\n\n"
        "Reenregistrez-le en UTF-8 ou en .xlsx depuis Excel."
    )


def list_sheets(path):
    """Noms des feuilles du classeur, pour le selecteur de l'interface.

    Un CSV n'a pas de feuille : la liste vide desactive le selecteur.
    """
    if is_csv(path):
        return []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()
    except Exception as exc:  # noqa: BLE001 - converti en message utilisateur
        raise _friendly_error(path, exc) from exc


def read_table(path, sheet_name=None):
    """Lit une feuille et retourne (dataframe, nom_de_feuille).

    dtype=object conserve les types d'origine : un entier reste un entier au
    lieu d'etre promu en float des qu'une cellule est vide (voir point 2).
    keep_default_na=False evite qu'un choix nomme "NA" soit pris pour un vide.
    """
    if not path:
        raise ExcelError("Aucun fichier selectionne.")
    if not os.path.exists(path):
        raise ExcelError(f"Fichier introuvable : {path}")

    if is_csv(path):
        encoding, separator = _sniff_csv(path)
        try:
            frame = pd.read_csv(
                path,
                sep=separator,
                dtype=object,
                keep_default_na=False,
                na_values=[""],
                encoding=encoding,
                skip_blank_lines=False,
            )
        except Exception as exc:  # noqa: BLE001 - converti en message utilisateur
            raise _friendly_error(path, exc) from exc
        target = ""
    else:
        target = sheet_name if sheet_name else 0
        try:
            frame = pd.read_excel(
                path,
                sheet_name=target,
                dtype=object,
                keep_default_na=False,
                na_values=[""],
            )
        except Exception as exc:  # noqa: BLE001 - converti en message utilisateur
            raise _friendly_error(path, exc) from exc

        if isinstance(frame, dict):  # securite : sheet_name=None renverrait un dict
            first = next(iter(frame))
            frame = frame[first]
            target = first

    frame.columns = [str(column).strip() for column in frame.columns]
    frame = _trim_trailing_blank_rows(frame)

    if is_csv(path):
        return frame, ""

    used = sheet_name if sheet_name else (list_sheets(path)[0] if target == 0 else target)
    return frame, used


def _trim_trailing_blank_rows(frame):
    """Retire les lignes vides de fin, jamais celles du milieu.

    Un classeur Excel traine souvent des lignes fantomes apres les donnees. Les
    supprimer toutes, y compris au milieu, decalerait la numerotation : la
    « ligne 42 » annoncee dans un rapport ne serait plus la ligne 42 du fichier
    de l'utilisateur, qui corrigerait alors la mauvaise ligne.
    """
    if frame.empty:
        return frame.reset_index(drop=True)

    non_blank = frame.notna().any(axis=1).to_numpy().nonzero()[0]
    if len(non_blank) == 0:
        return frame.iloc[0:0].reset_index(drop=True)
    return frame.iloc[: int(non_blank[-1]) + 1].reset_index(drop=True)


def file_signature(path):
    """Empreinte du contenu : identifie le fichier dans le registre (point 21).

    Fondee sur le contenu et non sur le chemin : renommer ou deplacer le fichier
    ne fait pas perdre l'historique, et une modification le distingue bien.
    """
    digest = hashlib.sha1()  # noqa: S324 - identification, pas de securite
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


# --------------------------------------------------------------------------
# Point 9 : modele Excel genere depuis le formulaire
# --------------------------------------------------------------------------

def _type_help(question):
    mapping = {
        "integer": "Nombre entier",
        "decimal": "Nombre decimal",
        "range": "Nombre",
        "date": "Date (AAAA-MM-JJ)",
        "datetime": "Date et heure",
        "time": "Heure (HH:MM:SS)",
        "text": "Texte libre",
        "select_one": "Un seul choix",
        "select_multiple": "Plusieurs choix separes par une virgule",
        "geopoint": "Latitude longitude",
        "barcode": "Code-barres",
        "calculate": "Valeur calculee (facultatif)",
        "acknowledge": "Case a cocher (true / false)",
    }
    return mapping.get(question.type, question.type)


def build_template(form_schema, target_path):
    """Ecrit un classeur pret a remplir pour ce formulaire."""
    questions = [
        question for question in form_schema.importable
        if not question.is_metadata and question.type != "calculate"
    ]
    if not questions:
        raise ExcelError("Ce formulaire ne contient aucune question importable.")

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = DATA_SHEET
    notice = workbook.create_sheet(NOTICE_SHEET)
    lists = workbook.create_sheet(LIST_SHEET)

    # --- feuille de saisie ------------------------------------------------
    for index, question in enumerate(questions, start=1):
        cell = data_sheet.cell(row=1, column=index, value=question.path)
        cell.font = _HEADER_FONT
        cell.fill = _REQUIRED_FILL if question.required else _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        note = f"{question.display_label()}\n\nType : {_type_help(question)}"
        if question.required:
            note += "\nReponse obligatoire"
        cell.comment = Comment(note, "Kobo Importer")

        width = min(42, max(14, len(question.path) + 4))
        data_sheet.column_dimensions[get_column_letter(index)].width = width

    data_sheet.freeze_panes = "A2"

    # --- listes de choix + validations ------------------------------------
    list_column = 0
    written_lists = {}
    for index, question in enumerate(questions, start=1):
        if not question.is_select or not question.choices:
            continue

        if question.list_name in written_lists:
            reference = written_lists[question.list_name]
        else:
            list_column += 1
            letter = get_column_letter(list_column)
            lists.cell(row=1, column=list_column, value=question.list_name or question.name)
            for offset, choice in enumerate(question.choices, start=2):
                lists.cell(row=offset, column=list_column, value=choice.name)
            reference = f"'{LIST_SHEET}'!${letter}$2:${letter}${len(question.choices) + 1}"
            written_lists[question.list_name or question.name] = reference

        if question.type == "select_one":
            # select_multiple accepte plusieurs valeurs : une liste fermee
            # empecherait la saisie, on se contente de documenter les choix.
            validation = DataValidation(type="list", formula1=reference, allow_blank=True)
            validation.error = "Choisissez une valeur dans la liste."
            validation.errorTitle = "Valeur non autorisee"
            data_sheet.add_data_validation(validation)
            letter = get_column_letter(index)
            validation.add(f"{letter}2:{letter}{TEMPLATE_ROWS}")

    lists.sheet_state = "hidden"

    # --- notice -----------------------------------------------------------
    titre = notice.cell(row=1, column=1, value=f"Modele pour : {form_schema.title}")
    titre.font = Font(bold=True, size=13)
    notice.cell(row=2, column=1, value=f"Formulaire {form_schema.uid} - version {form_schema.version}")
    notice.cell(
        row=3, column=1,
        value="Ne modifiez pas la ligne d'en-tete de la feuille « Donnees » : "
              "elle porte le nom technique des questions.",
    )

    headers = ["Colonne a remplir", "Question", "Type attendu", "Obligatoire", "Choix acceptes"]
    for column, label in enumerate(headers, start=1):
        cell = notice.cell(row=5, column=column, value=label)
        cell.font = Font(bold=True)
        cell.fill = _TITLE_FILL

    for offset, question in enumerate(questions, start=6):
        choices = ", ".join(choice.name for choice in question.choices) if question.choices else ""
        notice.cell(row=offset, column=1, value=question.path)
        notice.cell(row=offset, column=2, value=question.display_label())
        notice.cell(row=offset, column=3, value=_type_help(question))
        notice.cell(row=offset, column=4, value="Oui" if question.required else "")
        notice.cell(row=offset, column=5, value=choices)

    for column, width in enumerate((34, 46, 30, 12, 60), start=1):
        notice.column_dimensions[get_column_letter(column)].width = width
    notice.column_dimensions["E"].width = 60

    try:
        workbook.save(target_path)
    except OSError as exc:
        raise _friendly_write_error(target_path, exc) from exc
    return target_path


# --------------------------------------------------------------------------
# Point 11 : rapport d'erreurs reimportable
# --------------------------------------------------------------------------

def write_error_report(target_path, dataframe, failures, validation_report=None, context=None):
    """Ecrit un classeur des lignes non passees.

    failures : liste de dicts {row_index, status, message, http_status}
               ou row_index est la position (0..n-1) dans le dataframe lu.

    La feuille « A corriger » reprend les colonnes d'origine : apres correction,
    l'utilisateur reimporte ce fichier directement.
    """
    ordered = sorted(failures, key=lambda item: item.get("row_index", 0))
    positions = [
        item["row_index"] for item in ordered
        if isinstance(item.get("row_index"), int) and 0 <= item["row_index"] < len(dataframe)
    ]

    if positions:
        subset = dataframe.iloc[positions].copy()
    else:
        subset = dataframe.iloc[0:0].copy()

    # Le fichier traite peut deja etre un rapport corrige et reimporte : ses
    # deux colonnes de service sont alors remplacees, pas dupliquees.
    subset = subset.drop(columns=[REPORT_ROW_COLUMN, REPORT_REASON_COLUMN], errors="ignore")

    reasons = {item["row_index"]: item for item in ordered if isinstance(item.get("row_index"), int)}
    subset.insert(0, REPORT_ROW_COLUMN, [position + 2 for position in positions])
    subset.insert(1, REPORT_REASON_COLUMN, [
        _short_reason(reasons.get(position, {})) for position in positions
    ])

    try:
        _write_report_sheets(target_path, subset, ordered, validation_report, context)
    except OSError as exc:
        raise _friendly_write_error(target_path, exc) from exc

    _autosize(target_path)
    return target_path


def _write_report_sheets(target_path, subset, ordered, validation_report, context):
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        subset.to_excel(writer, sheet_name="A corriger", index=False)

        details = pd.DataFrame(
            [
                {
                    "Ligne Excel": item.get("row_index", -1) + 2,
                    "Statut": item.get("status", ""),
                    "Code HTTP": item.get("http_status") or "",
                    "Motif": item.get("message", ""),
                }
                for item in ordered
            ]
        )
        if details.empty:
            details = pd.DataFrame(columns=["Ligne Excel", "Statut", "Code HTTP", "Motif"])
        details.to_excel(writer, sheet_name="Details", index=False)

        if validation_report is not None and validation_report.row_issues:
            controls = pd.DataFrame(
                [
                    {
                        "Ligne Excel": issue.row_number,
                        "Colonne": issue.column,
                        "Valeur": issue.value,
                        "Motif": issue.message,
                    }
                    for issue in validation_report.row_issues
                ]
            )
            controls.to_excel(writer, sheet_name="Controle", index=False)

        if context:
            info = pd.DataFrame(list(context.items()), columns=["Information", "Valeur"])
            info.to_excel(writer, sheet_name="Contexte", index=False)


def _short_reason(item):
    message = str(item.get("message") or "").strip().replace("\n", " ")
    status = item.get("status") or ""
    if len(message) > 300:
        message = message[:297] + "..."
    return f"[{status}] {message}" if message else status


def _autosize(path, limit=60):
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        for column_cells in sheet.iter_cols(min_row=1, max_row=min(sheet.max_row, 200)):
            longest = 0
            letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                if cell.value is not None:
                    longest = max(longest, len(str(cell.value)))
            sheet.column_dimensions[letter].width = min(limit, max(12, longest + 2))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
    workbook.save(path)
